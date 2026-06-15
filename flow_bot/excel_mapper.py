from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
import xlrd

from .models import MappingResult, ProductRow

SCHEMA_KEYS = [
    "product_image",
    "product_name",
    "short_description",
    "long_description",
]

SCENE_SCHEMA_KEYS = [
    "scene_group_id",
    "scene_number",
    "scene_total",
    "scene_role",
    "scene_title",
    "scene_continuity_notes",
]

ALL_SCHEMA_KEYS = SCHEMA_KEYS + SCENE_SCHEMA_KEYS

ALIASES = {
    "product_image": [
        "product_image", "product_image_url", "image", "image_url", "img",
        "photo", "picture", "thumbnail", "main image", "gallery image",
        "link anh", "link ảnh", "url ảnh", "url hinh", "hình ảnh", "hinh anh",
        "ảnh sản phẩm", "anh san pham"
    ],
    "product_name": [
        "product_name", "tên sản phẩm", "ten san pham", "sản phẩm", "san pham",
        "name", "title", "product", "item name", "item"
    ],
    "short_description": [
        "short_description", "short desc", "short description", "brief",
        "summary", "tagline", "caption", "mô tả ngắn", "mo ta ngan",
        "mô tả sơ lược", "mo ta so luoc", "tóm tắt", "tom tat",
        "product_description", "description"
    ],
    "long_description": [
        "long_description", "long desc", "long description", "detail",
        "details", "full description", "body", "content", "mô tả dài",
        "mo ta dai", "mô tả chi tiết", "mo ta chi tiet", "chi tiết",
        "chi tiet", "nội dung", "noi dung", "selling_points",
        "điểm nổi bật", "diem noi bat", "lợi ích", "loi ich"
    ],
    "scene_group_id": [
        "scene_group_id", "scene group id", "scene group", "group id",
        "story group", "series id", "scene batch"
    ],
    "scene_number": [
        "scene_number", "scene number", "scene no", "scene index",
        "part number", "episode number", "segment number"
    ],
    "scene_total": [
        "scene_total", "scene total", "total scenes", "scene count",
        "total parts", "segment total"
    ],
    "scene_role": [
        "scene_role", "scene role", "role", "segment role", "video role"
    ],
    "scene_title": [
        "scene_title", "scene title", "title scene", "scene name",
        "segment title"
    ],
    "scene_continuity_notes": [
        "scene_continuity_notes", "scene continuity notes", "continuity notes",
        "continuity", "scene notes", "transition notes"
    ],
}


def normalize_text(value: str) -> str:
    source = (value or "").strip().lower()
    chars = {
        "àáạảãâầấậẩẫăằắặẳẵ": "a",
        "èéẹẻẽêềếệểễ": "e",
        "ìíịỉĩ": "i",
        "òóọỏõôồốộổỗơờớợởỡ": "o",
        "ùúụủũưừứựửữ": "u",
        "ỳýỵỷỹ": "y",
        "đ": "d",
    }
    for group, repl in chars.items():
        for char in group:
            source = source.replace(char, repl)
    return " ".join(source.split())


def tokenize(value: str) -> list[str]:
    cleaned = normalize_text(value)
    normalized = []
    token = []
    for char in cleaned:
        if char.isalnum():
            token.append(char)
        elif token:
            normalized.append("".join(token))
            token = []
    if token:
        normalized.append("".join(token))
    return normalized


def slugify(value: str) -> str:
    tokens = tokenize(value)
    slug = "-".join(tokens).strip("-")
    return slug or "single"


def score_header(header: str, aliases: Iterable[str], samples: list[str], schema_key: str) -> float:
    header_norm = normalize_text(header)
    header_tokens = set(tokenize(header))
    best = 0.0
    for alias in aliases:
        alias_norm = normalize_text(alias)
        alias_tokens = set(tokenize(alias))
        if header_norm == alias_norm:
            best = max(best, 1.0)
        elif alias_norm in header_norm or header_norm in alias_norm:
            best = max(best, 0.88)
        elif alias_tokens and header_tokens:
            overlap = len(alias_tokens & header_tokens)
            if overlap:
                best = max(best, min(0.84, overlap / max(len(alias_tokens), len(header_tokens)) + 0.35))

    non_empty = [sample for sample in samples if sample]
    avg_len = sum(len(sample) for sample in non_empty) / len(non_empty) if non_empty else 0
    image_count = sum(sample.startswith("http://") or sample.startswith("https://") for sample in non_empty)

    if schema_key == "product_image" and image_count:
        best = max(best, 0.97 if image_count >= max(1, len(non_empty) * 0.6) else 0.8)
    if schema_key == "product_name" and 0 < avg_len <= 90:
        best = max(best, 0.68)
    if schema_key == "short_description" and 0 < avg_len <= 220:
        best = max(best, 0.72)
    if schema_key == "long_description" and avg_len >= 80:
        best = max(best, 0.76)
    return round(best, 2)


def read_rows(file_path: Path, sheet_name: str | None = None) -> tuple[list[str], list[list[str]], str]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            rows = list(reader)
        if not rows:
            raise ValueError("CSV file is empty")
        headers = [str(cell).strip() for cell in rows[0]]
        data_rows = [[str(cell).strip() for cell in row] for row in rows[1:]]
        return headers, data_rows, "CSV"

    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        target_sheet = sheet_name or workbook.sheetnames[0]
        if target_sheet not in workbook.sheetnames:
            raise ValueError(f'Sheet "{target_sheet}" not found')
        sheet = workbook[target_sheet]
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            raise ValueError("Worksheet is empty")
        headers = [str(cell or "").strip() for cell in values[0]]
        data_rows = [[str(cell or "").strip() for cell in row] for row in values[1:]]
        return headers, data_rows, target_sheet

    if suffix == ".xls":
        workbook = xlrd.open_workbook(file_path.as_posix())
        target_sheet = sheet_name or workbook.sheet_names()[0]
        if target_sheet not in workbook.sheet_names():
            raise ValueError(f'Sheet "{target_sheet}" not found')
        sheet = workbook.sheet_by_name(target_sheet)
        if sheet.nrows == 0:
            raise ValueError("Worksheet is empty")
        headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
        data_rows = [
            [str(sheet.cell_value(row, col)).strip() for col in range(sheet.ncols)]
            for row in range(1, sheet.nrows)
        ]
        return headers, data_rows, target_sheet

    raise ValueError(f"Unsupported file type: {suffix}")


def auto_map_headers(headers: list[str], rows: list[list[str]]) -> MappingResult:
    sample_rows = rows[:5]
    mapped_headers: dict[str, str | None] = {}
    used_headers: set[str] = set()

    for schema_key in ALL_SCHEMA_KEYS:
        best_header = None
        best_score = 0.0
        for index, header in enumerate(headers):
            if not header or header in used_headers:
                continue
            samples = [str(row[index]).strip() if index < len(row) else "" for row in sample_rows]
            score = score_header(header, ALIASES[schema_key], samples, schema_key)
            if score > best_score:
                best_score = score
                best_header = header
        if best_header and best_score >= 0.55:
            mapped_headers[schema_key] = best_header
            used_headers.add(best_header)
        else:
            mapped_headers[schema_key] = None

    return MappingResult(mapped_headers=mapped_headers, row_count=len(rows))


def build_products(
    file_path: Path,
    sheet_name: str | None = None,
    mapping_override: dict[str, str | None] | None = None,
) -> tuple[list[ProductRow], MappingResult]:
    headers, rows, chosen_sheet = read_rows(file_path, sheet_name)
    mapping = auto_map_headers(headers, rows)
    if mapping_override:
        valid_headers = set(headers)
        cleaned_mapping: dict[str, str | None] = {}
        for schema_key in SCHEMA_KEYS:
            source_header = mapping_override.get(schema_key)
            cleaned_mapping[schema_key] = source_header if source_header in valid_headers else None
        for schema_key in SCENE_SCHEMA_KEYS:
            source_header = mapping.mapped_headers.get(schema_key)
            cleaned_mapping[schema_key] = source_header if source_header in valid_headers else None
        mapping.mapped_headers = cleaned_mapping
    mapping.sheet_name = chosen_sheet

    missing_required = [key for key in ("product_name", "short_description", "long_description") if not mapping.mapped_headers.get(key)]
    if missing_required:
        raise ValueError(f"Missing required column mapping: {', '.join(missing_required)}")

    products: list[ProductRow] = []
    missing_scene_total_flags: list[bool] = []
    header_index = {header: idx for idx, header in enumerate(headers)}
    for raw_row in rows:
        product_data: dict[str, str] = {}
        for schema_key in ALL_SCHEMA_KEYS:
            source_header = mapping.mapped_headers.get(schema_key)
            value = ""
            if source_header is not None:
                idx = header_index[source_header]
                value = str(raw_row[idx]).strip() if idx < len(raw_row) else ""
            product_data[schema_key] = value
        if not product_data["short_description"] and product_data["long_description"]:
            product_data["short_description"] = product_data["long_description"][:180].strip()
        if not product_data["long_description"] and product_data["short_description"]:
            product_data["long_description"] = product_data["short_description"]
        scene_group_id = (product_data.get("scene_group_id") or "").strip()
        scene_title = product_data.get("scene_title") or ""
        scene_role = product_data.get("scene_role") or "single"
        missing_scene_total = not str(product_data.get("scene_total") or "").strip()
        try:
            scene_number = max(1, int(float(product_data.get("scene_number") or "1")))
        except ValueError:
            scene_number = 1
        try:
            scene_total = max(1, int(float(product_data.get("scene_total") or "1")))
        except ValueError:
            scene_total = 1
        product = ProductRow(
            product_image=product_data["product_image"],
            product_name=product_data["product_name"],
            short_description=product_data["short_description"],
            long_description=product_data["long_description"],
            scene_group_id=scene_group_id,
            scene_number=scene_number,
            scene_total=scene_total,
            scene_role=scene_role or "single",
            scene_title=scene_title,
            scene_continuity_notes=product_data.get("scene_continuity_notes") or "",
        )
        if any((product.product_name, product.short_description, product.long_description)):
            product.validate()
            if not product.scene_group_id:
                product.scene_group_id = slugify(product.product_name)
            if not product.scene_title:
                product.scene_title = product.product_name
            products.append(product)
            missing_scene_total_flags.append(missing_scene_total)

    group_counts: dict[str, int] = {}
    for product in products:
        group_id = product.scene_group_id or slugify(product.product_name)
        group_counts[group_id] = group_counts.get(group_id, 0) + 1

    for index, product in enumerate(products):
        if missing_scene_total_flags[index]:
            product.scene_total = group_counts.get(product.scene_group_id, product.scene_total)

    mapping.row_count = len(products)
    return products, mapping
